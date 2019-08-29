from openpyxl import load_workbook
import pyautogui as auto
from time import sleep

form_position = (225, 372)
wb = load_workbook('contacts.xlsx')
def main():  
    sheets = wb.sheetnames
    for all_sheet in sheets:
        sheet = wb[all_sheet]
        sleep(5)
        auto.click(form_position, clicks=2, button='left')
        for row in range(1, sheet.max_row + 1):
                for column in "ABCD":
                        cell_value = f"{column}{row}"
                        cell = str(sheet[cell_value].value)
                        auto.typewrite(cell)                     
                        auto.press('tab')                       
if __name__ == "__main__":
    main()